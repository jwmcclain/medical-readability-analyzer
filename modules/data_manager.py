"""
Data management module
Handles Excel export with multiple sheets and formatting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, List
import os
import logging
import sys

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def export_to_excel(results: List[Dict], stats: Dict, search_term: str, output_path: str) -> str:
    """
    Export all data to formatted Excel workbook.
    
    Args:
        results: List of result dictionaries
        stats: Statistical analysis results
        search_term: Original search term
        output_path: Path to save Excel file
    
    Returns:
        Path to saved file
    """
    logger.info(f"Exporting data to Excel: {output_path}")
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Summary
        create_summary_sheet(writer, df, stats, search_term)
        
        # Sheet 2: Readability Data
        create_data_sheet(writer, df)
        
        # Sheet 3: Statistical Analysis
        create_stats_sheet(writer, stats)
        
        # Sheet 4: Full Text
        create_text_sheet(writer, df)
        
        # Sheet 5: Classification Details
        create_classification_sheet(writer, df)
    
    # Apply formatting
    apply_excel_formatting(output_path)
    
    logger.info(f"Excel file saved: {output_path}")
    return output_path


def create_summary_sheet(writer, df: pd.DataFrame, stats: Dict, search_term: str):
    """Create summary information sheet."""
    summary_data = {
        'Metric': [
            'Search Term',
            'Analysis Date',
            'Total URLs Analyzed',
            'Institutional Sources',
            'Private Sources',
            'Mean GFI Score',
            'Mean SMOG Score',
            'Mean FKG Score',
            'Mean ARI Score',
            'Overall Mean Readability',
            'URLs at Universal Level (≤8th grade)',
        ],
        'Value': [
            search_term,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            len(df),
            len(df[df['source_type'] == 'Institutional']) if 'source_type' in df.columns else 0,
            len(df[df['source_type'] == 'Private']) if 'source_type' in df.columns else 0,
            f"{df['GFI'].mean():.2f}" if 'GFI' in df.columns else 'N/A',
            f"{df['SMOG'].mean():.2f}" if 'SMOG' in df.columns else 'N/A',
            f"{df['FKG'].mean():.2f}" if 'FKG' in df.columns else 'N/A',
            f"{df['ARI'].mean():.2f}" if 'ARI' in df.columns else 'N/A',
            f"{df['mean_readability'].mean():.2f}" if 'mean_readability' in df.columns else 'N/A',
            len(df[df['mean_readability'] <= 8]) if 'mean_readability' in df.columns else 0,
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)


def create_data_sheet(writer, df: pd.DataFrame):
    """Create main readability data sheet."""
    # Select relevant columns
    columns = ['rank', 'url', 'domain', 'source_type', 'classification_confidence',
               'GFI', 'SMOG', 'FKG', 'ARI', 'mean_readability', 
               'word_count', 'sentence_count', 'status']
    
    # Filter to available columns
    available_cols = [col for col in columns if col in df.columns]
    data_df = df[available_cols].copy()
    
    # Round numeric columns
    numeric_cols = ['GFI', 'SMOG', 'FKG', 'ARI', 'mean_readability']
    for col in numeric_cols:
        if col in data_df.columns:
            data_df[col] = data_df[col].round(1)
    
    data_df.to_excel(writer, sheet_name='Readability_Data', index=False)


def create_stats_sheet(writer, stats: Dict):
    """Create statistical analysis results sheet."""
    if not stats:
        pd.DataFrame({'Note': ['No statistical analysis performed']}).to_excel(
            writer, sheet_name='Statistical_Analysis', index=False
        )
        return
    
    rows = []
    
    # Descriptive statistics
    rows.append(['DESCRIPTIVE STATISTICS BY SOURCE TYPE', '', '', ''])
    rows.append(['Metric', 'Source Type', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'N'])
    
    for metric, source_stats in stats.get('descriptive', {}).items():
        for source_type, stats_dict in source_stats.items():
            rows.append([
                metric,
                source_type,
                f"{stats_dict['mean']:.2f}",
                f"{stats_dict['median']:.2f}",
                f"{stats_dict['std']:.2f}",
                f"{stats_dict['min']:.2f}",
                f"{stats_dict['max']:.2f}",
                stats_dict['n']
            ])
    
    rows.append(['', '', '', '', '', '', ''])
    
    # Group comparisons
    rows.append(['GROUP COMPARISONS (Institutional vs. Private)', '', '', ''])
    rows.append(['Metric', 'Test Used', 'P-value', 'Significant', 'Effect Size'])
    
    for metric, comparison in stats.get('comparisons', {}).items():
        if 'error' not in comparison:
            rows.append([
                metric,
                comparison.get('test_used', ''),
                f"{comparison.get('p_value', 0):.4f}",
                '***' if comparison.get('significant') else 'No',
                f"{comparison.get('effect_size', 0):.3f}"
            ])
    
    stats_df = pd.DataFrame(rows)
    stats_df.to_excel(writer, sheet_name='Statistical_Analysis', index=False, header=False)


def create_text_sheet(writer, df: pd.DataFrame):
    """Create sheet with full text content."""
    if 'content' in df.columns:
        text_df = df[['rank', 'url', 'content']].copy()
        text_df.to_excel(writer, sheet_name='Full_Text', index=False)
    else:
        pd.DataFrame({'Note': ['No content available']}).to_excel(
            writer, sheet_name='Full_Text', index=False
        )


def create_classification_sheet(writer, df: pd.DataFrame):
    """Create classification details sheet."""
    if 'source_type' in df.columns:
        class_cols = ['url', 'domain', 'source_type', 'classification_confidence']
        available_cols = [col for col in class_cols if col in df.columns]
        class_df = df[available_cols].copy()
        class_df.to_excel(writer, sheet_name='Classification_Details', index=False)
    else:
        pd.DataFrame({'Note': ['No classification data available']}).to_excel(
            writer, sheet_name='Classification_Details', index=False
        )


def apply_excel_formatting(filepath: str):
    """Apply formatting to Excel workbook."""
    wb = openpyxl.load_workbook(filepath)
    
    # Format Readability_Data sheet
    if 'Readability_Data' in wb.sheetnames:
        ws = wb['Readability_Data']
        apply_conditional_formatting(ws)
        format_headers(ws)
        auto_adjust_columns(ws)
    
    # Format other sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        format_headers(ws)
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    wb.save(filepath)


def apply_conditional_formatting(ws):
    """Apply color coding to readability scores."""
    from openpyxl.formatting.rule import CellIsRule
    
    # Find mean_readability column
    header_row = 1
    mean_col = None
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value == 'mean_readability':
            mean_col = col
            break
    
    if mean_col:
        col_letter = openpyxl.utils.get_column_letter(mean_col)
        
        # Green for ≤ 8 (universal)
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='lessThanOrEqual', formula=['8'], fill=green_fill)
        )
        
        # Yellow for 8-10 (acceptable)
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='between', formula=['8', '10'], fill=yellow_fill)
        )
        
        # Red for > 10 (difficult)
        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='greaterThan', formula=['10'], fill=red_fill)
        )


def format_headers(ws):
    """Format header row."""
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_adjust_columns(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    # Test export
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    
    test_results = [
        {
            'rank': 1,
            'url': 'https://www.example.com',
            'domain': 'example.com',
            'source_type': 'Institutional',
            'classification_confidence': 3,
            'GFI': 10.5,
            'SMOG': 11.2,
            'FKG': 9.8,
            'ARI': 10.1,
            'mean_readability': 10.4,
            'word_count': 500,
            'sentence_count': 25,
            'status': 'success',
            'content': 'Sample content here...'
        }
    ]
    
    test_stats = {
        'descriptive': {
            'GFI': {
                'Institutional': {'mean': 10.0, 'median': 10.5, 'std': 2.0, 'min': 7.0, 'max': 15.0, 'n': 50}
            }
        }
    }
    
    output = 'test_output.xlsx'
    export_to_excel(test_results, test_stats, 'test', output)
    print(f"Test Excel file created: {output}")
